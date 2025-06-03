import streamlit as st
import pandas as pd
import datetime
from pathlib import Path
import plotly.express as px
import plotly.graph_objects as go
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from fpdf import FPDF
import io
import time
import base64
import openpyxl
import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload

# Configure page
st.set_page_config(
    page_title="Excel Practice Test",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        text-align: center;
        color: #0078d4;
        border-bottom: 3px solid #0078d4;
        padding-bottom: 20px;
        margin-bottom: 30px;
    }
    .question-box {
        background: #f8f9fa;
        padding: 15px;
        border-left: 3px solid #28a745;
        border-radius: 4px;
        margin: 15px 0;
    }
    .instructions-box {
        background: #fff3cd;
        border: 1px solid #ffeaa7;
        padding: 15px;
        border-radius: 4px;
        margin: 20px 0;
    }
    .data-table {
        font-size: 12px;
    }
    .metric-card {
        background: white;
        padding: 20px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .timer {
        font-size: 24px;
        font-weight: bold;
        color: #dc3545;
        text-align: center;
        margin: 10px 0;
    }
    .pivot-input {
        margin: 5px 0;
    }
</style>
""", unsafe_allow_html=True)

# Load secrets
ADMIN_PASSWORD = st.secrets.get("admin_password", "admin123")
ADMIN_EMAILS = st.secrets.get("admin_emails", ["admin1@example.com", "admin2@example.com"]).split(",")
EMAIL_SENDER = st.secrets.get("email_sender", "your_email@example.com")
EMAIL_PASSWORD = st.secrets.get("email_password", "your_email_password")
SMTP_SERVER = st.secrets.get("smtp_server", "smtp.gmail.com")
SMTP_PORT = st.secrets.get("smtp_port", 587)
GOOGLE_SHEET_URL = st.secrets.get("GOOGLE_SHEET_URL", "your-google-sheet-url")
DRIVE_FOLDER_ID = st.secrets.get("DRIVE_FOLDER_ID", "your-drive-folder-id")

# Initialize Google Sheets and Drive API
scopes = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive"
]
creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
sheets_client = gspread.authorize(creds)
drive_service = build("drive", "v3", credentials=creds)

# Open Google Sheet
try:
    sheet = sheets_client.open_by_url(GOOGLE_SHEET_URL).sheet1
except Exception as e:
    st.error(f"Failed to connect to Google Sheets: {str(e)}")
    st.stop()

# Initialize session state
if 'user_answers' not in st.session_state:
    st.session_state.user_answers = {}
if 'user_info' not in st.session_state:
    st.session_state.user_info = {}
if 'test_submitted' not in st.session_state:
    st.session_state.test_submitted = False
if 'time_remaining' not in st.session_state:
    st.session_state.time_remaining = 30 * 60  # 30 minutes in seconds
if 'timer_active' not in st.session_state:
    st.session_state.timer_active = False
if 'shuffled_questions' not in st.session_state:
    st.session_state.shuffled_questions = []

# Employee data
employee_data = [
    {"Employee": "Saravana Kumar R", "Gender": "Male", "Marital Status": "Unmarried", "Region": "South", "Location": "Trichy", "Department": "TSG & IT Hardware", "Total Amount Due": 2000},
    {"Employee": "Narsi Ram Meena", "Gender": "Male", "Marital Status": "Married", "Region": "North", "Location": "Lucknow", "Department": "TSG & IT Hardware", "Total Amount Due": 6000},
    {"Employee": "Shahbaz Khan", "Gender": "Male", "Marital Status": "Married", "Region": "North", "Location": "Agra", "Department": "Customer Service Division", "Total Amount Due": 4400},
    {"Employee": "Aman Mishra", "Gender": "Male", "Marital Status": "Unmarried", "Region": "West", "Location": "Satara", "Department": "TSG & IT Hardware", "Total Amount Due": 2300},
    {"Employee": "Bherulal Sharma", "Gender": "Male", "Marital Status": "Unmarried", "Region": "West", "Location": "Satara", "Department": "Accounts", "Total Amount Due": 10000},
    {"Employee": "Brajesh Sharma", "Gender": "Male", "Marital Status": "Married", "Region": "North", "Location": "Lucknow", "Department": "TSG & IT Hardware", "Total Amount Due": 15000},
    {"Employee": "Suraj Mahor", "Gender": "Male", "Marital Status": "Unmarried", "Region": "North", "Location": "Lucknow", "Department": "TSG & IT Hardware", "Total Amount Due": 14000},
    {"Employee": "Shikha Yadav", "Gender": "Female", "Marital Status": "Married", "Region": "East", "Location": "Noida", "Department": "Sales", "Total Amount Due": 200},
    {"Employee": "Sunita Gautam Dudhe", "Gender": "Female", "Marital Status": "Married", "Region": "West", "Location": "Nagpur", "Department": "Customer Service Division", "Total Amount Due": 123},
    {"Employee": "Dhan Das", "Gender": "Male", "Marital Status": "Unmarried", "Region": "East", "Location": "Guwahati", "Department": "TSG & IT Hardware", "Total Amount Due": 0},
    {"Employee": "Anamika Singh Chaudhary", "Gender": "Female", "Marital Status": "Unmarried", "Region": "North", "Location": "Agra", "Department": "Customer Service Division", "Total Amount Due": 1},
    {"Employee": "Chaitram Dhanraj Shahu", "Gender": "Male", "Marital Status": "Married", "Region": "West", "Location": "Nagpur", "Department": "Customer Service Division", "Total Amount Due": 12300},
    {"Employee": "Dev Singh Saharawat", "Gender": "Male", "Marital Status": "Unmarried", "Region": "North", "Location": "Ambala", "Department": "Accounts", "Total Amount Due": 5300},
    {"Employee": "Santosh Kumar Singh", "Gender": "Male", "Marital Status": "Married", "Region": "North", "Location": "Noida", "Department": "TSG & IT Hardware", "Total Amount Due": 47000},
]

# Correct answers for MCQs only
correct_answers = {
    "q1": "a",  # True
    "q2": "b",  # Column heading
    "q3": "b",  # Conditional Formatting
    "q4": "a",  # Right
    "q5": "b",  # Does not change
    "q6": "b",  # IF
    "q7": "a",  # True
    "q8": "a",  # Show only rows where Category = "Food"
}

# Image for Question 8
QUESTION_8_IMAGE = "https://raw.githubusercontent.com/MrSingh529/excel-practice-test/main/images/pivot_table_slicer.png"

def upload_to_drive(file_data, filename, folder_id):
    """Upload a file to Google Drive and return its shareable link"""
    try:
        file_metadata = {
            "name": filename,
            "parents": [folder_id]
        }
        media = MediaIoBaseUpload(io.BytesIO(file_data), mimetype="image/jpeg")
        file = drive_service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id"
        ).execute()
        
        # Make the file publicly accessible
        drive_service.permissions().create(
            fileId=file["id"],
            body={"role": "reader", "type": "anyone"}
        ).execute()
        
        # Get the shareable link
        file_link = f"https://drive.google.com/file/d/{file['id']}/view"
        return file_link
    except Exception as e:
        st.error(f"Failed to upload to Google Drive: {str(e)}")
        return None

def load_submissions():
    """Load submissions from Google Sheets"""
    try:
        records = sheet.get_all_records()
        submissions = []
        for record in records:
            answers = {}
            for key in record:
                if key.startswith("Q") and key.endswith("Screenshot URL"):
                    answers[key.lower().replace(" ", "_")] = record[key]
                elif key.startswith("Q"):
                    answers[key.lower()] = record[key]
            submissions.append({
                "timestamp": record["Timestamp"],
                "user_info": {
                    "name": record["Name"],
                    "employee_id": record["Employee ID"],
                    "department": record["Department"],
                    "email": record["Email"]
                },
                "score": int(record["MCQ Score"].split("/")[0]),
                "total": int(record["MCQ Score"].split("/")[1]),
                "percentage": float(record["Percentage"].replace("%", "")),
                "answers": answers
            })
        return submissions
    except Exception as e:
        st.error(f"Failed to load submissions: {str(e)}")
        return []

def save_submission(submission):
    """Save a new submission to Google Sheets"""
    try:
        row = [
            submission["timestamp"],
            submission["user_info"]["name"],
            submission["user_info"]["employee_id"],
            submission["user_info"]["department"],
            submission["user_info"]["email"],
            f"{submission['score']}/{submission['total']}",
            f"{submission['percentage']:.1f}%",
            "PASS" if submission["percentage"] >= 70 else "FAIL",
            submission["answers"].get("q1", ""),
            submission["answers"].get("q2", ""),
            submission["answers"].get("q3", ""),
            submission["answers"].get("q4", ""),
            submission["answers"].get("q5", ""),
            submission["answers"].get("q6", ""),
            submission["answers"].get("q7", ""),
            submission["answers"].get("q8", ""),
            submission["answers"].get("q9a_screenshot_url", ""),
            submission["answers"].get("q9b_screenshot_url", ""),
            submission["answers"].get("q10_screenshot_url", "")
        ]
        sheet.append_row(row)
    except Exception as e:
        st.error(f"Failed to save submission: {str(e)}")

def calculate_score(user_answers):
    """Calculate test score for MCQs only"""
    score = 0
    total = len(correct_answers)  # 8 MCQs
    for q_id, correct_answer in correct_answers.items():
        if user_answers.get(q_id) == correct_answer:
            score += 1
    return score, total

def send_email(recipient, subject, body):
    """Send email notification"""
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_SENDER
        msg['To'] = recipient
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        st.error(f"Failed to send email: {str(e)}")
        return False

def generate_certificate(name, score, total, date):
    """Generate PDF certificate"""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 24)
    pdf.cell(0, 20, "Certificate of Achievement", ln=True, align="C")
    pdf.set_font("Arial", "", 16)
    pdf.ln(20)
    pdf.cell(0, 10, f"This certifies that", ln=True, align="C")
    pdf.set_font("Arial", "B", 20)
    pdf.cell(0, 10, name, ln=True, align="C")
    pdf.set_font("Arial", "", 16)
    pdf.ln(10)
    pdf.cell(0, 10, f"has successfully completed the Excel Practice Test", ln=True, align="C")
    pdf.cell(0, 10, f"Score: {score}/{total} (MCQs only)", ln=True, align="C")
    pdf.cell(0, 10, f"Date: {date}", ln=True, align="C")
    pdf.ln(10)
    pdf.cell(0, 10, "Note: PivotTable questions (9 & 10) are graded separately by admins.", ln=True, align="C")
    pdf.ln(20)
    pdf.set_font("Arial", "I", 12)
    pdf.cell(0, 10, "Learning & Development Department", ln=True, align="C")

    # Save PDF to bytes
    output = io.BytesIO()
    pdf_output = pdf.output(dest='S').encode('latin1')
    output.write(pdf_output)
    output.seek(0)
    return output

def create_detailed_analytics(submissions):
    """Create detailed analytics for admin"""
    df = pd.DataFrame([{
        "timestamp": s["timestamp"],
        "name": s["user_info"]["name"],
        "department": s["user_info"]["department"],
        "score": s["score"],
        "total": s["total"],
        "percentage": s["percentage"],
        "answers": s["answers"]
    } for s in submissions])
    
    if df.empty:
        return None, None, None, None
    
    # Question-wise accuracy for MCQs
    question_accuracy = {}
    for q_id in correct_answers.keys():
        correct_count = sum(1 for _, row in df.iterrows() if row["answers"].get(q_id) == correct_answers[q_id])
        total_attempts = len(df)
        question_accuracy[q_id] = (correct_count / total_attempts * 100) if total_attempts > 0 else 0
    
    # Performance over time
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    performance_over_time = df.groupby(df["timestamp"].dt.date)["percentage"].mean().reset_index()
    
    # Department-wise performance
    dept_performance = df.groupby("department")["percentage"].agg(["mean", "count"]).reset_index()
    
    # Individual question analysis
    question_details = []
    for q_id in correct_answers.keys():
        answers = df["answers"].apply(lambda x: x.get(q_id, "Not answered")).value_counts()
        question_details.append({"Question": q_id, "Answer Distribution": answers.to_dict()})
    
    return question_accuracy, performance_over_time, dept_performance, question_details

# Timer logic
def update_timer():
    if st.session_state.timer_active and st.session_state.time_remaining > 0:
        time.sleep(1)
        st.session_state.time_remaining -= 1
    if st.session_state.time_remaining <= 0:
        st.session_state.timer_active = False
        st.session_state.test_submitted = True
        st.rerun()

# Sidebar navigation
st.sidebar.title("Navigation")
page = st.sidebar.selectbox("Choose a page:", 
    ["🏠 Home", "📝 Take Test", "👨‍💼 Admin Dashboard"])

if page == "🏠 Home":
    st.markdown('<h1 class="main-header">📊 Excel Practice Test</h1>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; color: #666; font-style: italic;">Learning & Development Department: Together we learn, together we soar.</p>', unsafe_allow_html=True)
    
    st.markdown("""
    ## Welcome to the Digital Excel Practice Test!
    
    This comprehensive test evaluates your Excel knowledge across multiple areas:
    
    ### 📋 Test Sections:
    - **Section A**: Multiple Choice Questions (8 questions)
    - **Section B**: Data Analysis using Employee Dataset  
    - **Section C**: PivotTable Understanding (2 questions)
    
    ### 🎯 Learning Objectives:
    - Master Excel fundamentals
    - Understand data analysis concepts
    - Learn PivotTable functionality
    - Practice conditional formatting
    - Explore logical functions
    
    ### 📊 Features:
    - Interactive online test with timer
    - Instant score calculation and email notifications
    - Progress tracking
    - Admin analytics
    - Certificate generation for passing
    """)
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info("**Total Questions**: 10 (8 MCQs + 2 PivotTables)")
    with col2:
        st.success("**Time Limit**: 30 minutes")
    with col3:
        st.warning("**Passing Score**: 70% (MCQs only)")

elif page == "📝 Take Test":
    st.markdown('<h1 class="main-header">📝 Excel Practice Test</h1>', unsafe_allow_html=True)
    
    if not st.session_state.test_submitted:
        # User information form
        with st.expander("📋 Enter Your Information", expanded=True):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("👤 Full Name*", value=st.session_state.user_info.get('name', ''))
                employee_id = st.text_input("🆔 Employee ID*", value=st.session_state.user_info.get('employee_id', ''))
            with col2:
                department = st.selectbox("🏢 Department*", 
                    ["", "TSG & IT Hardware", "Customer Service Division", "Accounts", "Sales", "HR", "Other"],
                    index=0 if not st.session_state.user_info.get('department') else 
                    ["", "TSG & IT Hardware", "Customer Service Division", "Accounts", "Sales", "HR", "Other"].index(st.session_state.user_info.get('department')))
                email = st.text_input("📧 Email*", value=st.session_state.user_info.get('email', ''))
        
        # Store user info
        st.session_state.user_info = {
            'name': name,
            'employee_id': employee_id,
            'department': department,
            'email': email
        }
        
        # Instructions
        st.markdown("""
        <div class="instructions-box">
        <strong>📝 Instructions:</strong><br>
        • Answer all 10 questions (8 multiple-choice and 2 PivotTable questions)<br>
        • For multiple-choice, select the best answer<br>
        • For PivotTable questions (9 & 10), download the Employee Data as an Excel file, create the PivotTables in Excel, and upload screenshots of your PivotTables (max 5 MB each)<br>
        • PivotTable questions will be graded manually by admins<br>
        • Review the employee data table for context<br>
        • Submit your answers within 30 minutes<br>
        • You can change answers before final submission
        </div>
        """, unsafe_allow_html=True)
        
        # Timer display
        st.markdown(f"""
        <div class="timer">
        ⏰ Time Remaining: {int(st.session_state.time_remaining // 60)}:{int(st.session_state.time_remaining % 60):02d}
        </div>
        """, unsafe_allow_html=True)
        
        # Start timer on first interaction
        if not st.session_state.timer_active:
            st.session_state.timer_active = True
        
        # Update timer every second
        st_autorefresh = st.empty()
        update_timer()
        
        # Employee Data Display
        st.markdown("## Section B: Employee Data Reference")
        st.markdown("*Use this data to understand the context for the questions below:*")
        
        df = pd.DataFrame(employee_data)
        st.dataframe(df, use_container_width=True)
        
        # Download Employee Data as Excel
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False, engine='openpyxl')
        excel_buffer.seek(0)
        st.download_button(
            label="📥 Download Employee Data as Excel",
            data=excel_buffer,
            file_name="employee_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        # Questions
        st.markdown("## Section A: Multiple Choice Questions")
        
        questions = [
            {
                "id": "q1",
                "text": "The AutoSum feature adds up the numbers in a column or row that you specify.",
                "options": ["True", "False"]
            },
            {
                "id": "q2", 
                "text": 'In Excel, the label "AAA" (as seen at the top of a worksheet) is an example of a:',
                "options": ["Cell reference", "Column heading", "Name box entry", "Row heading"]
            },
            {
                "id": "q3",
                "text": "__________ quickly highlight important information in a spreadsheet by applying formatting options such as data bars, color scales, or icon sets.",
                "options": ["Cell references", "Conditional Formatting", "Excel tables", "PivotTables"]
            },
            {
                "id": "q4",
                "text": "As a rule, Excel will __________-align numbers in a cell.",
                "options": ["Right", "Left", "Top", "Bottom"]
            },
            {
                "id": "q5",
                "text": "When you copy a formula that contains an absolute reference (e.g., $A$1) to a new location, the absolute reference:",
                "options": ["Updates automatically to reflect the new row/column", "Does not change", "Becomes bold", "Gets a dotted outline in its cell"]
            },
            {
                "id": "q6",
                "text": "Which of the following is a logical function in Excel?",
                "options": ["AVERAGE", "IF", "SUMPRODUCT", "VLOOKUP"]
            },
            {
                "id": "q7",
                "text": "PivotTables are a powerful tool used to quickly group, summarize, and rearrange larger datasets.",
                "options": ["True", "False"]
            },
            {
                "id": "q8",
                "text": 'Consider a PivotTable with a slicer connected to the "Category" field. If you click "Food" on that slicer, the PivotTable will:',
                "options": ['Show only rows where Category = "Food"', 'Show all rows except those where Category = "Food"', 'Not change (slicer has no effect)'],
                "image": QUESTION_8_IMAGE
            }
        ]
        
        # Display questions in fixed order (no shuffling)
        for i, question in enumerate(questions, 1):
            st.markdown(f"""
            <div class="question-box">
            <strong>Question {i}:</strong> {question['text']}
            </div>
            """, unsafe_allow_html=True)
            
            # Display image for Question 8
            if "image" in question:
                st.image(question["image"], caption="PivotTable Slicer Example", use_column_width=True)
            
            option_labels = [chr(97 + j) for j in range(len(question['options']))]  # a, b, c, d
            formatted_options = [f"{label}. {option}" for label, option in zip(option_labels, question['options'])]
            
            selected = st.radio(
                f"Select your answer for Question {i}:",
                options=option_labels,
                format_func=lambda x: formatted_options[ord(x) - 97],
                key=question['id'],
                index=None
            )
            
            if selected:
                st.session_state.user_answers[question['id']] = selected
        
        # PivotTable Questions
        st.markdown("## Section C: PivotTable Questions")
        st.markdown("**Note**: These questions require you to create PivotTables in Excel using the downloaded Employee Data file. Please upload screenshots of your PivotTables below (max 5 MB each). These will be reviewed manually by admins.")
        
        # Question 9
        st.markdown("""
        <div class="question-box">
        <strong>Question 9:</strong> Using the Employee Data table above, create two PivotTables:<br>
        a. A PivotTable that shows, for each Region, the total of "Total Amount Due"<br>
        b. A PivotTable that shows, for each Department, the total of "Total Amount Due"
        </div>
        """, unsafe_allow_html=True)
        
        # Question 9a: Upload screenshot
        st.markdown("**9a. Total Amount Due by Region**")
        q9a_screenshot = st.file_uploader("Upload a screenshot of your PivotTable for 9a (PNG/JPG, max 5 MB)", type=["png", "jpg", "jpeg"], key="q9a_screenshot")
        if q9a_screenshot:
            # Check file size (5 MB = 5 * 1024 * 1024 bytes)
            file_data = q9a_screenshot.read()
            if len(file_data) > 5 * 1024 * 1024:
                st.error("File size exceeds 5 MB limit. Please upload a smaller file.")
            else:
                # Upload to Google Drive
                filename = f"{name}_{employee_id}_q9a_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                q9a_screenshot_url = upload_to_drive(file_data, filename, DRIVE_FOLDER_ID)
                if q9a_screenshot_url:
                    st.session_state.user_answers["q9a_screenshot_url"] = q9a_screenshot_url
                    st.image(file_data, caption="Uploaded PivotTable for 9a", use_column_width=True)
        
        # Question 9b: Upload screenshot
        st.markdown("**9b. Total Amount Due by Department**")
        q9b_screenshot = st.file_uploader("Upload a screenshot of your PivotTable for 9b (PNG/JPG, max 5 MB)", type=["png", "jpg", "jpeg"], key="q9b_screenshot")
        if q9b_screenshot:
            file_data = q9b_screenshot.read()
            if len(file_data) > 5 * 1024 * 1024:
                st.error("File size exceeds 5 MB limit. Please upload a smaller file.")
            else:
                filename = f"{name}_{employee_id}_q9b_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                q9b_screenshot_url = upload_to_drive(file_data, filename, DRIVE_FOLDER_ID)
                if q9b_screenshot_url:
                    st.session_state.user_answers["q9b_screenshot_url"] = q9b_screenshot_url
                    st.image(file_data, caption="Uploaded PivotTable for 9b", use_column_width=True)
        
        # Question 10
        st.markdown("""
        <div class="question-box">
        <strong>Question 10:</strong> Using the Employee Data table above, build a PivotTable in a new worksheet that shows, for each Region, the count of employees by Gender.
        </div>
        """, unsafe_allow_html=True)
        
        q10_screenshot = st.file_uploader("Upload a screenshot of your PivotTable for Question 10 (PNG/JPG, max 5 MB)", type=["png", "jpg", "jpeg"], key="q10_screenshot")
        if q10_screenshot:
            file_data = q10_screenshot.read()
            if len(file_data) > 5 * 1024 * 1024:
                st.error("File size exceeds 5 MB limit. Please upload a smaller file.")
            else:
                filename = f"{name}_{employee_id}_q10_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
                q10_screenshot_url = upload_to_drive(file_data, filename, DRIVE_FOLDER_ID)
                if q10_screenshot_url:
                    st.session_state.user_answers["q10_screenshot_url"] = q10_screenshot_url
                    st.image(file_data, caption="Uploaded PivotTable for Question 10", use_column_width=True)
        
        # Submit button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.button("🚀 Submit Test", type="primary", use_container_width=True):
                # Validate user info
                if not all([name, employee_id, department, email]):
                    st.error("⚠️ Please fill in all required information fields!")
                elif len({k: v for k, v in st.session_state.user_answers.items() if k.startswith('q') and not k.endswith('_screenshot_url')}) < len(correct_answers):
                    st.error(f"⚠️ Please answer all multiple-choice questions! You have answered {len({k: v for k, v in st.session_state.user_answers.items() if k.startswith('q') and not k.endswith('_screenshot_url')})} out of {len(correct_answers)} MCQs.")
                elif not all(st.session_state.user_answers.get(key) for key in ["q9a_screenshot_url", "q9b_screenshot_url", "q10_screenshot_url"]):
                    st.error("⚠️ Please upload screenshots for all PivotTable questions (9a, 9b, and 10)!")
                else:
                    # Calculate score for MCQs only
                    score, total = calculate_score(st.session_state.user_answers)
                    percentage = (score / total) * 100
                    
                    # Create submission record
                    submission = {
                        "timestamp": datetime.datetime.now().isoformat(),
                        "user_info": st.session_state.user_info,
                        "answers": st.session_state.user_answers,
                        "score": score,
                        "total": total,
                        "percentage": percentage
                    }
                    
                    # Save submission to Google Sheets
                    save_submission(submission)
                    
                    # Send email to user
                    user_body = f"""
                    Dear {name},
                    
                    Thank you for completing the Excel Practice Test.
                    Your MCQ Score: {score}/{total} ({percentage:.1f}%)
                    Status: {'PASS' if percentage >= 70 else 'NEEDS IMPROVEMENT'} (MCQs only)
                    Note: Your PivotTable submissions (Questions 9 & 10) will be reviewed by admins.
                    
                    Regards,
                    Learning & Development Department
                    """
                    send_email(email, "Excel Practice Test Results", user_body)
                    
                    # Send email to admins
                    admin_body = f"""
                    New Test Submission:
                    Name: {name}
                    Employee ID: {employee_id}
                    Department: {department}
                    MCQ Score: {score}/{total} ({percentage:.1f}%)
                    Status: {'PASS' if percentage >= 70 else 'NEEDS IMPROVEMENT'} (MCQs only)
                    Note: Please review the PivotTable screenshots for Questions 9 & 10 in the Admin Dashboard.
                    """
                    for admin_email in ADMIN_EMAILS:
                        send_email(admin_email.strip(), "New Excel Test Submission", admin_body)
                    
                    st.session_state.test_submitted = True
                    st.session_state.timer_active = False
                    st.rerun()
    
    else:
        # Show results
        score, total = calculate_score(st.session_state.user_answers)
        percentage = (score / total) * 100
        
        st.success("🎉 Test Submitted Successfully!")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("MCQ Score", f"{score}/{total}")
        with col2:
            st.metric("Percentage", f"{percentage:.1f}%")
        with col3:
            status = "PASS" if percentage >= 70 else "NEEDS IMPROVEMENT"
            color = "green" if percentage >= 70 else "red"
            st.metric("Result", status)
        
        st.info("Note: Your PivotTable submissions (Questions 9 & 10) will be reviewed by admins separately.")
        
        # Certificate generation for passing users (MCQs only)
        if percentage >= 70:
            date = datetime.datetime.now().strftime("%Y-%m-%d")
            cert_buffer = generate_certificate(st.session_state.user_info["name"], score, total, date)
            st.download_button(
                label="📜 Download Certificate (MCQs only)",
                data=cert_buffer,
                file_name=f"Excel_Practice_Certificate_{st.session_state.user_info['name']}.pdf",
                mime="application/pdf"
            )
        
        # Detailed results for MCQs
        st.markdown("## 📊 Detailed Results (MCQs)")
        results_data = []
        for i, (q_id, correct_answer) in enumerate(correct_answers.items(), 1):
            user_answer = st.session_state.user_answers.get(q_id, "Not answered")
            is_correct = user_answer == correct_answer
            user_answer_display = user_answer.upper() if user_answer != "Not answered" else user_answer
            results_data.append({
                "Question": i,
                "Your Answer": user_answer_display,
                "Correct Answer": correct_answer.upper(),
                "Result": "✅ Correct" if is_correct else "❌ Incorrect"
            })
        
        results_df = pd.DataFrame(results_data)
        st.dataframe(results_df, use_container_width=True)
        
        if st.button("🔄 Take Test Again"):
            st.session_state.user_answers = {}
            st.session_state.user_info = {}
            st.session_state.test_submitted = False
            st.session_state.time_remaining = 30 * 60
            st.session_state.timer_active = False
            st.session_state.shuffled_questions = []
            st.rerun()

elif page == "👨‍💼 Admin Dashboard":
    st.markdown('<h1 class="main-header">👨‍💼 Admin Dashboard</h1>', unsafe_allow_html=True)
    
    # Admin authentication
    if 'admin_authenticated' not in st.session_state:
        st.session_state.admin_authenticated = False
    
    if not st.session_state.admin_authenticated:
        st.warning("🔐 Admin access required")
        password = st.text_input("Enter admin password:", type="password")
        if st.button("Login"):
            if password == ADMIN_PASSWORD:
                st.session_state.admin_authenticated = True
                st.success("✅ Admin access granted!")
                st.rerun()
            else:
                st.error("❌ Invalid password!")
    else:
        submissions = load_submissions()
        
        if not submissions:
            st.info("📝 No test submissions yet.")
        else:
            # Summary statistics
            total_submissions = len(submissions)
            avg_score = sum(s['percentage'] for s in submissions) / total_submissions
            pass_rate = sum(1 for s in submissions if s['percentage'] >= 70) / total_submissions * 100
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Submissions", total_submissions)
            with col2:
                st.metric("Average MCQ Score", f"{avg_score:.1f}%")
            with col3:
                st.metric("Pass Rate (MCQs)", f"{pass_rate:.1f}%")
            
            # Detailed Analytics for MCQs
            question_accuracy, performance_over_time, dept_performance, question_details = create_detailed_analytics(submissions)
            
            st.subheader("📈 Detailed Analytics (MCQs)")
            
            col1, col2 = st.columns(2)
            with col1:
                # Question Accuracy
                fig_accuracy = px.bar(x=list(question_accuracy.keys()), y=list(question_accuracy.values()),
                                    title="Question-wise Accuracy (%)",
                                    labels={'x': 'Question', 'y': 'Accuracy %'})
                st.plotly_chart(fig_accuracy, use_container_width=True)
                
                # Performance Over Time
                fig_time = px.line(performance_over_time, x="timestamp", y="percentage",
                                 title="Performance Over Time",
                                 labels={'timestamp': 'Date', 'percentage': 'Average Score %'})
                st.plotly_chart(fig_time, use_container_width=True)
            
            with col2:
                # Department Performance
                fig_dept = px.bar(dept_performance, x="department", y="mean",
                                title="Department-wise Performance",
                                labels={'department': 'Department', 'mean': 'Average Score %'},
                                text="count")
                fig_dept.update_traces(textposition='auto')
                st.plotly_chart(fig_dept, use_container_width=True)
            
            # Question Details
            st.subheader("🔍 Question Analysis (MCQs)")
            for detail in question_details:
                st.write(f"**{detail['Question'].upper()} Answer Distribution:**")
                dist_df = pd.DataFrame.from_dict(detail["Answer Distribution"], orient="index", columns=["Count"])
                st.dataframe(dist_df, use_container_width=True)
            
            # Detailed submissions table with PivotTable screenshot links
            st.subheader("📋 All Submissions")
            
            display_data = []
            for s in submissions:
                row = {
                    "Timestamp": s['timestamp'][:19].replace('T', ' '),
                    "Name": s['user_info']['name'],
                    "Employee ID": s['user_info']['employee_id'],
                    "Department": s['user_info']['department'],
                    "Email": s['user_info']['email'],
                    "MCQ Score": f"{s['score']}/{s['total']}",
                    "Percentage": f"{s['percentage']:.1f}%",
                    "Status": "PASS" if s['percentage'] >= 70 else "FAIL",
                    "Q9a Screenshot": "View",
                    "Q9b Screenshot": "View",
                    "Q10 Screenshot": "View"
                }
                display_data.append(row)
            
            submissions_df = pd.DataFrame(display_data)
            
            # Display the table with clickable buttons to view screenshots
            for idx, row in submissions_df.iterrows():
                col1, col2, col3, col4, col5, col6, col7, col8, col9, col10, col11 = st.columns([2, 2, 2, 2, 2, 1, 1, 1, 1, 1, 1])
                with col1:
                    st.write(row["Timestamp"])
                with col2:
                    st.write(row["Name"])
                with col3:
                    st.write(row["Employee ID"])
                with col4:
                    st.write(row["Department"])
                with col5:
                    st.write(row["Email"])
                with col6:
                    st.write(row["MCQ Score"])
                with col7:
                    st.write(row["Percentage"])
                with col8:
                    st.write(row["Status"])
                with col9:
                    if st.button("View Q9a", key=f"q9a_{idx}"):
                        url = submissions[idx]["answers"].get("q9a_screenshot_url")
                        if url:
                            st.markdown(f"[View Q9a PivotTable]({url})")
                        else:
                            st.warning("No screenshot uploaded.")
                with col10:
                    if st.button("View Q9b", key=f"q9b_{idx}"):
                        url = submissions[idx]["answers"].get("q9b_screenshot_url")
                        if url:
                            st.markdown(f"[View Q9b PivotTable]({url})")
                        else:
                            st.warning("No screenshot uploaded.")
                with col11:
                    if st.button("View Q10", key=f"q10_{idx}"):
                        url = submissions[idx]["answers"].get("q10_screenshot_url")
                        if url:
                            st.markdown(f"[View Q10 PivotTable]({url})")
                        else:
                            st.warning("No screenshot uploaded.")
            
            # Download submissions as Excel
            if st.button("📥 Download All Submissions (Excel)"):
                excel_data = []
                for s in submissions:
                    row = {
                        "Timestamp": s['timestamp'][:19].replace('T', ' '),
                        "Name": s['user_info']['name'],
                        "Employee ID": s['user_info']['employee_id'],
                        "Department": s['user_info']['department'],
                        "Email": s['user_info']['email'],
                        "MCQ Score": f"{s['score']}/{s['total']}",
                        "Percentage": f"{s['percentage']:.1f}%",
                        "Status": "PASS" if s['percentage'] >= 70 else "FAIL",
                        "Q1": s['answers'].get("q1", ""),
                        "Q2": s['answers'].get("q2", ""),
                        "Q3": s['answers'].get("q3", ""),
                        "Q4": s['answers'].get("q4", ""),
                        "Q5": s['answers'].get("q5", ""),
                        "Q6": s['answers'].get("q6", ""),
                        "Q7": s['answers'].get("q7", ""),
                        "Q8": s['answers'].get("q8", ""),
                        "Q9a Screenshot URL": s['answers'].get("q9a_screenshot_url", ""),
                        "Q9b Screenshot URL": s['answers'].get("q9b_screenshot_url", ""),
                        "Q10 Screenshot URL": s['answers'].get("q10_screenshot_url", "")
                    }
                    excel_data.append(row)
                
                df = pd.DataFrame(excel_data)
                excel_buffer = io.BytesIO()
                df.to_excel(excel_buffer, index=False, engine='openpyxl')
                excel_buffer.seek(0)
                st.download_button(
                    label="Download Submissions as Excel",
                    data=excel_buffer,
                    file_name=f"excel_test_submissions_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        if st.button("🚪 Admin Logout"):
            st.session_state.admin_authenticated = False
            st.rerun()

# Footer
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: #666;'>"
    "📊 Excel Practice Test | Learning & Development Department<br>"
    "Together we learn, together we soar 🚀"
    "</div>", 
    unsafe_allow_html=True
)